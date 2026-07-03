"""
excel_builder.py
Builds a rich, color-coded Excel workbook from matched job data.
4 sheets: Today's Jobs | Skills Gap | Watchlist | Weekly Summary (Mondays)
"""
import logging
from collections import Counter
from datetime import datetime, date
from pathlib import Path
from typing import List, Dict, Optional

logger = logging.getLogger(__name__)

# ─── Colour palette ───────────────────────────────────────────
GREEN_FILL  = "C6EFCE"   # 80-100% match
YELLOW_FILL = "FFEB9C"   # 60-79%
ORANGE_FILL = "FFDAB9"   # 40-59%
WHITE_FILL  = "FFFFFF"   # 30-39%
HEADER_FILL = "1F3864"   # Dark navy header
HEADER_FONT = "FFFFFF"   # White text on header

URGENCY_COLORS = {
    "HIGH":   "FF4444",
    "MEDIUM": "FFA500",
    "LOW":    "228B22",
}


def _get_fill(pct: int):
    from openpyxl.styles import PatternFill
    if pct >= 80:
        hex_color = GREEN_FILL
    elif pct >= 60:
        hex_color = YELLOW_FILL
    elif pct >= 40:
        hex_color = ORANGE_FILL
    else:
        hex_color = WHITE_FILL
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _header_style():
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    font = Font(bold=True, color=HEADER_FONT, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return fill, font, align


def _urgency_icon(urgency: str) -> str:
    return {"HIGH": "🔴 Apply Today", "MEDIUM": "🟡 This Week", "LOW": "🟢 This Month"}.get(urgency, "⚪")


MAIN_COLUMNS = [
    ("🏢 Company",         20),
    ("💼 Role",            30),
    ("🎯 Match %",         12),
    ("📄 Profile Used",    18),
    ("🌟 Urgency",         18),
    ("🔗 Apply Link",      40),
    ("📍 Location",        18),
    ("💰 Salary/Stipend",  20),
    ("🏗️ Company Type",   16),
    ("🏢 Size",            12),
    ("✅ Matched Skills",  35),
    ("❌ Missing Skills",  35),
    ("💬 Why Fit",         45),
    ("📅 Posted Date",     14),
    ("📌 Source",          14),
    ("✍️ Applied?",        12),
]


def _write_main_sheet(ws, jobs: List[Dict]):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Header row
    h_fill, h_font, h_align = _header_style()
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 35
    for col_idx, (col_name, col_width) in enumerate(MAIN_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Data rows
    for row_idx, job in enumerate(jobs, start=2):
        pct = job.get("match_percentage", 0)
        row_fill = _get_fill(pct)
        urgency = job.get("urgency", "LOW")

        values = [
            job.get("company", ""),
            job.get("title", ""),
            f"{pct}%",
            job.get("matched_profile", ""),
            _urgency_icon(urgency),
            job.get("apply_url", ""),
            job.get("location", ""),
            job.get("salary", ""),
            job.get("company_type", ""),
            job.get("company_size", ""),
            ", ".join(job.get("matched_skills", [])[:8]),
            ", ".join(job.get("missing_skills", [])[:8]),
            job.get("why_good_fit", ""),
            str(job.get("posted_date", ""))[:10],
            job.get("source", ""),
            "",  # Applied? — user fills in
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Hyperlink on Apply Link column
            if col_idx == 6 and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
            else:
                cell.font = Font(size=10)

        # Match % column — bold + larger
        pct_cell = ws.cell(row=row_idx, column=3)
        pct_cell.font = Font(bold=True, size=11)
        pct_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 28

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MAIN_COLUMNS))}1"


def _write_skills_gap_sheet(ws, jobs: List[Dict]):
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    h_fill, h_font, h_align = _header_style()

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40

    headers = ["Missing Skill", "Jobs Affected", "Suggested Resource"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align

    # Tally all missing skills
    all_missing = []
    for job in jobs:
        all_missing.extend(job.get("missing_skills", []))

    counter = Counter(all_missing)
    resources = {
        "docker": "docs.docker.com / FreeCodeCamp",
        "kubernetes": "kubernetes.io/docs",
        "aws":    "AWS Free Tier + CloudQuest",
        "react":  "react.dev (official)",
        "typescript": "typescriptlang.org",
        "system design": "Gaurav Sen YouTube",
        "sql":    "Mode Analytics SQL Tutorial",
        "machine learning": "fast.ai / Coursera ML",
        "deep learning": "fast.ai Part 1",
        "golang": "go.dev/tour",
        "rust":   "rustlings on GitHub",
    }

    for row_idx, (skill, count) in enumerate(counter.most_common(20), start=2):
        resource = next((v for k, v in resources.items() if k in skill.lower()), "Search Coursera / YouTube")
        ws.cell(row=row_idx, column=1, value=skill).font = Font(size=10)
        ws.cell(row=row_idx, column=2, value=count).font = Font(bold=True, size=10)
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=resource).font = Font(size=10, italic=True)

    # Bar chart
    if counter:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Top Missing Skills This Week"
        chart.y_axis.title = "Skill"
        chart.x_axis.title = "Jobs Count"
        chart.style = 10
        chart.width = 20
        chart.height = 12

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=min(len(counter)+1, 16))
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=min(len(counter)+1, 16))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "E2")


def _write_watchlist_sheet(ws, jobs: List[Dict], watchlist: List[str]):
    from openpyxl.styles import Font, Alignment, PatternFill

    h_fill, h_font, h_align = _header_style()
    from openpyxl.utils import get_column_letter

    cols = [("🏢 Company", 22), ("💼 Role", 28), ("🎯 Match %", 12), ("🔗 Apply", 40), ("📅 Posted", 14)]
    for col_idx, (name, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    watchlist_lower = [w.lower() for w in watchlist]
    wl_jobs = [j for j in jobs if j.get("company", "").lower() in watchlist_lower]

    if not wl_jobs:
        ws.cell(row=2, column=1, value="No jobs found from your watchlist companies today.")
        return

    for row_idx, job in enumerate(wl_jobs, start=2):
        pct = job.get("match_percentage", 0)
        ws.cell(row=row_idx, column=1, value=job.get("company", ""))
        ws.cell(row=row_idx, column=2, value=job.get("title", ""))
        ws.cell(row=row_idx, column=3, value=f"{pct}%").font = Font(bold=True)
        link_cell = ws.cell(row=row_idx, column=4, value=job.get("apply_url", ""))
        if job.get("apply_url"):
            link_cell.hyperlink = job["apply_url"]
            link_cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=row_idx, column=5, value=str(job.get("posted_date", ""))[:10])


def _write_summary_sheet(ws, jobs: List[Dict], config_profile_name: str, tips: List[str]):
    from openpyxl.styles import Font, Alignment

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 55

    today = datetime.now().strftime("%B %d, %Y")
    ws.cell(row=1, column=1, value="📊 DAILY REPORT SUMMARY").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated: {today}").font = Font(italic=True)

    stats = [
        ("Total jobs found today", len(jobs)),
        ("80–100% match (🟢 Apply Now)", sum(1 for j in jobs if j.get("match_percentage", 0) >= 80)),
        ("60–79% match (🟡 Strong)",     sum(1 for j in jobs if 60 <= j.get("match_percentage", 0) < 80)),
        ("40–59% match (🟠 Consider)",   sum(1 for j in jobs if 40 <= j.get("match_percentage", 0) < 60)),
        ("Urgent (Apply Today 🔴)",       sum(1 for j in jobs if j.get("urgency") == "HIGH")),
        ("Top source",                   Counter([j.get("source","") for j in jobs]).most_common(1)[0][0] if jobs else "-"),
        ("Best match job",               max(jobs, key=lambda x: x.get("match_percentage",0)).get("title","") if jobs else "-"),
        ("Best match company",           max(jobs, key=lambda x: x.get("match_percentage",0)).get("company","") if jobs else "-"),
        ("Best match %",                 f"{max((j.get('match_percentage',0) for j in jobs), default=0)}%"),
    ]

    for row_idx, (label, value) in enumerate(stats, start=4):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=str(value))

    if tips:
        ws.cell(row=14, column=1, value="🧠 RESUME IMPROVEMENT TIPS").font = Font(bold=True, size=12)
        for i, tip in enumerate(tips, start=15):
            ws.cell(row=i, column=1, value=f"💡 Tip {i-14}")
            ws.cell(row=i, column=2, value=tip)


# ─────────────────────────────────────────────────────────────────
# Main builder function
# ─────────────────────────────────────────────────────────────────

def build_excel(
    jobs: List[Dict],
    watchlist: List[str],
    profile_name: str,
    tips: List[str],
    output_path: Path,
) -> Path:
    """
    Build and save the full Excel workbook.
    Returns the path to the saved file.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        raise

    wb = Workbook()

    # Sheet 1: Today's Jobs
    ws_main = wb.active
    ws_main.title = "📋 Today's Jobs"
    # Sort: first by urgency (HIGH first), then by match %
    urgency_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    sorted_jobs = sorted(
        jobs,
        key=lambda j: (urgency_order.get(j.get("urgency", "LOW"), 2), -j.get("match_percentage", 0))
    )
    _write_main_sheet(ws_main, sorted_jobs)

    # Sheet 2: Skills Gap
    ws_gap = wb.create_sheet("📉 Skills Gap")
    _write_skills_gap_sheet(ws_gap, jobs)

    # Sheet 3: Watchlist
    ws_wl = wb.create_sheet("🏆 Watchlist")
    _write_watchlist_sheet(ws_wl, jobs, watchlist)

    # Sheet 4: Summary
    ws_sum = wb.create_sheet("📊 Summary")
    _write_summary_sheet(ws_sum, jobs, profile_name, tips)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"[Excel] Saved to {output_path}")
    return output_path
